import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
import re

def main():
    st.set_page_config(
        page_title="残業時間集計アプリ",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("📊 残業時間集計アプリ")
    st.markdown("---")
    
    # タブの作成
    tab1, tab2 = st.tabs(["📈 残業時間集計", "📅 休日・平日仕訳"])
    
    with tab1:
        overtime_tab()
    
    with tab2:
        holiday_tab()

def overtime_tab():
    """残業時間集計タブの内容"""
    st.header("📈 残業時間集計")
    
    # ファイルアップロード
    uploaded_file = st.file_uploader(
        "エクセルファイルをアップロードしてください",
        type=['xlsx', 'xls'],
        help="複数のシートを持つエクセルファイルをアップロードしてください"
    )
    
    if uploaded_file is not None:
        try:
            # エクセルファイルを読み込み（data_only=Trueで計算結果を取得）
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            sheet_names = workbook.sheetnames
            
            st.success(f"ファイルが正常に読み込まれました。シート数: {len(sheet_names)}")
            
            # 固定シートの確認
            fixed_sheets = ["まとめ", "記入例", "報告書format", "残業代"]
            member_sheets = [sheet for sheet in sheet_names if sheet not in fixed_sheets]
            
            st.info(f"固定シート: {fixed_sheets}")
            st.info(f"メンバーシート: {member_sheets}")
            
            if member_sheets:
                # 残業時間の集計
                overtime_data = extract_overtime_data(workbook, member_sheets)
                
                if overtime_data:
                    display_results(overtime_data)
                else:
                    st.warning("残業時間のデータが見つかりませんでした。")
            else:
                st.warning("メンバーのシートが見つかりませんでした。")
                
        except Exception as e:
            st.error(f"ファイルの読み込み中にエラーが発生しました: {str(e)}")

def holiday_tab():
    """休日・平日仕訳タブの内容"""
    st.header("📅 休日・平日仕訳")
    
    # ファイルアップロード
    uploaded_file = st.file_uploader(
        "エクセルファイルをアップロードしてください",
        type=['xlsx', 'xls'],
        help="複数のシートを持つエクセルファイルをアップロードしてください",
        key="holiday_uploader"
    )
    
    if uploaded_file is not None:
        try:
            # エクセルファイルを読み込み（data_only=Trueで計算結果を取得）
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            sheet_names = workbook.sheetnames
            
            st.success(f"ファイルが正常に読み込まれました。シート数: {len(sheet_names)}")
            
            # 固定シートの確認
            fixed_sheets = ["まとめ", "記入例", "報告書format", "残業代"]
            member_sheets = [sheet for sheet in sheet_names if sheet not in fixed_sheets]
            
            st.info(f"固定シート: {fixed_sheets}")
            st.info(f"メンバーシート: {member_sheets}")
            
            if member_sheets:
                # 休日・平日仕訳の集計
                holiday_data = extract_holiday_data(workbook, member_sheets)
                
                if holiday_data:
                    display_holiday_results(holiday_data)
                    
                    # 残業代シートから単価を読み込み
                    overtime_rates = read_overtime_sheet(workbook)
                    
                    if overtime_rates:
                        # 残業代を計算
                        pay_data = calculate_overtime_pay(holiday_data, overtime_rates)
                        
                        if pay_data:
                            display_overtime_pay_results(pay_data, holiday_data)
                        else:
                            st.warning("残業代の計算に失敗しました。")
                    else:
                        st.warning("残業代シートから単価データを読み込めませんでした。")
                else:
                    st.warning("休日・平日仕訳のデータが見つかりませんでした。")
            else:
                st.warning("メンバーのシートが見つかりませんでした。")
                
        except Exception as e:
            st.error(f"ファイルの読み込み中にエラーが発生しました: {str(e)}")

def extract_overtime_data(workbook, member_sheets):
    """残業時間データを抽出する"""
    overtime_data = {}
    
    # 時間帯の定義
    time_slots = {
        'K39': '休日時間帯の応動（09:00-18:00）',
        'O39': '平日・休日時間外の応動（18:00-22:00）',
        'S39': '平日・休日深夜の応動（22:00-05:00）',
        'W39': '平日・休日時間外の応動（05:00-09:00）'
    }
    
    for sheet_name in member_sheets:
        try:
            worksheet = workbook[sheet_name]
            member_data = {}
            
            for cell_ref, time_slot in time_slots.items():
                # セルK39, O39, S39, W39の値を取得
                cell_value = worksheet[cell_ref].value
                
                # 結合セルの場合、下のセル（K40, O40, S40, W40）も確認
                if cell_value is None:
                    # 結合セルの下のセルを確認
                    next_cell_ref = cell_ref.replace('39', '40')
                    cell_value = worksheet[next_cell_ref].value
                
                if cell_value is not None:
                    # 表示用の形式と集計用の数値を両方保存
                    display_format = parse_time_to_display_format(cell_value)
                    time_hours = parse_time_to_hours(cell_value)
                    
                    if time_hours > 0:
                        member_data[time_slot] = {
                            'display': display_format,
                            'hours': time_hours
                        }
                    else:
                        member_data[time_slot] = {
                            'display': "",  # 空白セル
                            'hours': 0
                        }
                else:
                    member_data[time_slot] = {
                        'display': "",  # 空白セル
                        'hours': 0
                    }
            
            # 全メンバーを追加（データがなくても表示）
            overtime_data[sheet_name] = member_data
                
        except Exception as e:
            st.warning(f"シート '{sheet_name}' の処理中にエラーが発生しました: {str(e)}")
            continue
    
    return overtime_data

def parse_time_to_display_format(time_value):
    """時間値を表示用の形式に変換する（1:30形式）"""
    if time_value is None:
        return ""  # 空白セル
    
    # datetime.timeオブジェクトの場合
    if hasattr(time_value, 'hour') and hasattr(time_value, 'minute'):
        hours = time_value.hour
        minutes = time_value.minute
        result = f"{hours}:{minutes:02d}"
        print(f"DEBUG: datetime.time {time_value} -> {result}")
        return result
    
    # 文字列の場合
    time_str = str(time_value).strip()
    if not time_str or time_str == '':
        return ""  # 空白セル
    
    # 時間:分:秒の形式をパース（例: "1:30:00" -> "1:30"）
    if ':' in time_str:
        try:
            parts = time_str.split(':')
            if len(parts) >= 2:
                hours = int(parts[0])
                minutes = int(parts[1])
                # 0:00の場合は空白を返す
                if hours == 0 and minutes == 0:
                    return ""  # 空白セル
                result = f"{hours}:{minutes:02d}"
                print(f"DEBUG: 時間文字列 {time_str} -> {result}")
                return result
        except Exception as e:
            print(f"DEBUG: パースエラー {time_str}: {e}")
            pass
    
    # 数値の場合（エクセルの時間値は小数で表現される）
    try:
        # エクセルの時間値は1日=1.0で表現されるので、24倍して時間に変換
        if isinstance(time_value, (int, float)):
            total_hours = time_value * 24
            hours = int(total_hours)
            minutes = int((total_hours - hours) * 60)
            # 0:00の場合は空白を返す
            if hours == 0 and minutes == 0:
                return ""  # 空白セル
            result = f"{hours}:{minutes:02d}"
            print(f"DEBUG: エクセル時間値 {time_value} -> {result}")
            return result
        else:
            # 数値として認識された場合
            total_hours = float(time_str)
            hours = int(total_hours)
            minutes = int((total_hours - hours) * 60)
            # 0:00の場合は空白を返す
            if hours == 0 and minutes == 0:
                return ""  # 空白セル
            result = f"{hours}:{minutes:02d}"
            print(f"DEBUG: 数値として認識 {time_str} -> {result}")
            return result
    except:
        # 文字列から数値を抽出
        import re
        numbers = re.findall(r'\d+\.?\d*', time_str)
        if numbers:
            total_hours = float(numbers[0])
            hours = int(total_hours)
            minutes = int((total_hours - hours) * 60)
            # 0:00の場合は空白を返す
            if hours == 0 and minutes == 0:
                return ""  # 空白セル
            result = f"{hours}:{minutes:02d}"
            print(f"DEBUG: 文字列から数値抽出 {time_str} -> {result}")
            return result
        print(f"DEBUG: 認識できない形式 {time_str}")
        return ""  # 空白セル

def parse_time_to_hours(time_value):
    """時間値を時間数に変換する（集計用）"""
    if time_value is None:
        return 0
    
    # datetime.timeオブジェクトの場合
    if hasattr(time_value, 'hour') and hasattr(time_value, 'minute'):
        hours = time_value.hour
        minutes = time_value.minute
        result = hours + minutes / 60
        return result
    
    # datetime.datetimeオブジェクトの場合
    if hasattr(time_value, 'date') and hasattr(time_value, 'time'):
        # 日付部分を除いて時間部分のみを取得
        time_part = time_value.time()
        hours = time_part.hour
        minutes = time_part.minute
        result = hours + minutes / 60
        return result
    
    # 文字列の場合
    time_str = str(time_value).strip()
    if not time_str or time_str == '':
        return 0
    
    # 時間:分:秒の形式をパース
    if ':' in time_str:
        try:
            parts = time_str.split(':')
            if len(parts) >= 2:
                hours = int(parts[0])
                minutes = int(parts[1])
                result = hours + minutes / 60
                return result
        except Exception as e:
            pass
    
    # 数値の場合（エクセルの時間値は小数で表現される）
    try:
        if isinstance(time_value, (int, float)):
            # エクセルの時間値は1日=1.0で表現されるので、24倍して時間に変換
            result = time_value * 24
            return result
        else:
            # 文字列を数値として変換
            result = float(time_str)
            # 1未満の場合は時間値として扱う（1日=1.0）
            if result < 1:
                result = result * 24
            return result
    except Exception as e:
        import re
        numbers = re.findall(r'\d+\.?\d*', time_str)
        if numbers:
            result = float(numbers[0])
            # 1未満の場合は時間値として扱う
            if result < 1:
                result = result * 24
            return result
        return 0

def extract_holiday_data(workbook, member_sheets):
    """休日・平日仕訳データを抽出する"""
    holiday_data = {}
    
    # 時間帯の定義
    time_slots = {
        'K': '休日時間帯の応動（09:00-18:00）',
        'O': '平日・休日時間外の応動（18:00-22:00）',
        'S': '平日・休日深夜の応動（22:00-05:00）',
        'W': '平日・休日時間外の応動（05:00-09:00）'
    }
    
    for sheet_name in member_sheets:
        try:
            worksheet = workbook[sheet_name]
            member_data = {}
            
            for column, time_slot in time_slots.items():
                holiday_hours = 0
                weekday_hours = 0
                
                # 8行目から38行目までチェック
                for row in range(8, 39):
                    # 時間セル（K8, O8, S8, W8など）
                    time_cell = f"{column}{row}"
                    time_value = worksheet[time_cell].value
                    
                    # 時間が00:01以上の場合のみ処理
                    if time_value is not None:
                        time_hours = parse_time_to_hours(time_value)
                        # 00:01以上（約0.000694時間以上）の場合のみ処理
                        if time_hours > 0.000694:  # 1分 = 1/60/24 = 0.000694時間
                            # B列の曜日情報を取得（DATE関数の結果を取得）
                            day_cell = f"B{row}"
                            day_value = worksheet[day_cell].value
                            
                            # C列の祝日情報を取得
                            holiday_cell = f"C{row}"
                            holiday_value = worksheet[holiday_cell].value
                            
                            # 休日・平日の判定
                            is_holiday = is_holiday_day(day_value, holiday_value)
                            
                            if is_holiday:
                                holiday_hours += time_hours
                            else:
                                weekday_hours += time_hours
                
                member_data[time_slot] = {
                    'holiday_hours': holiday_hours,
                    'weekday_hours': weekday_hours,
                    'total_hours': holiday_hours + weekday_hours
                }
            
            holiday_data[sheet_name] = member_data
                
        except Exception as e:
            st.warning(f"シート '{sheet_name}' の処理中にエラーが発生しました: {str(e)}")
            continue
    
    return holiday_data

def is_holiday_day(day_value, holiday_value):
    """曜日と祝日情報から休日かどうかを判定する"""
    if day_value is None:
        return False
    
    # エクセルの日付シリアル値（整数）の場合
    if isinstance(day_value, int):
        # エクセルの日付シリアル値をdatetimeオブジェクトに変換
        # エクセルの基準日は1900年1月1日（ただし、1900年は閏年として扱われるバグがある）
        from datetime import datetime, timedelta
        try:
            # エクセルの基準日（1900年1月1日）から日数を加算
            base_date = datetime(1899, 12, 30)  # エクセルの基準日
            target_date = base_date + timedelta(days=day_value)
            weekday = target_date.weekday()
            
            # 土曜日(5)と日曜日(6)は休日
            if weekday in [5, 6]:
                return True
            
            # 月〜金の場合、C列に「祝日」と記載がある場合は休日
            if holiday_value is not None and str(holiday_value).strip() == '祝日':
                return True
            
            return False
        except Exception as e:
            # エラーの場合は祝日情報で判定
            if holiday_value is not None and str(holiday_value).strip() == '祝日':
                return True
            return False
    
    # DATE関数の結果（datetimeオブジェクト）の場合
    if hasattr(day_value, 'weekday'):
        # weekday()は月曜日=0, 日曜日=6
        weekday = day_value.weekday()
        
        # 土曜日(5)と日曜日(6)は休日
        if weekday in [5, 6]:
            return True
        
        # 月〜金の場合、C列に「祝日」と記載がある場合は休日
        if holiday_value is not None and str(holiday_value).strip() == '祝日':
            return True
        
        return False
    
    # 文字列の場合
    day_str = str(day_value).strip()
    
    # 土日は休日
    if day_str in ['土', '日']:
        return True
    
    # 月〜金の場合、C列に「祝日」と記載がある場合は休日
    if day_str in ['月', '火', '水', '木', '金']:
        if holiday_value is not None and str(holiday_value).strip() == '祝日':
            return True
        return False
    
    return False

def display_holiday_results(holiday_data):
    """休日・平日仕訳結果を表示する"""
    st.markdown("## 📅 休日・平日仕訳結果")
    
    # データフレームを作成（指定された形式）
    df_data = []
    for member, data in holiday_data.items():
        row = {'メンバー': member}
        
        # 各時間帯の休日・平日時間を追加
        time_slots = [
            '休日時間帯の応動（09:00-18:00）',
            '平日・休日時間外の応動（18:00-22:00）',
            '平日・休日深夜の応動（22:00-05:00）',
            '平日・休日時間外の応動（05:00-09:00）'
        ]
        
        for time_slot in time_slots:
            if time_slot in data:
                time_data = data[time_slot]
                # 休日時間
                row[f'{time_slot}_休日'] = format_hours(time_data['holiday_hours'])
                # 平日時間
                row[f'{time_slot}_平日'] = format_hours(time_data['weekday_hours'])
            else:
                row[f'{time_slot}_休日'] = ""
                row[f'{time_slot}_平日'] = ""
        
        df_data.append(row)
    
    if df_data:
        df = pd.DataFrame(df_data)
        
        # 表示
        st.dataframe(df, use_container_width=True)
        
        # ダウンロードボタン
        csv = df.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📥 CSVファイルとしてダウンロード",
            data=csv,
            file_name=f"休日平日仕訳_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
        
        # 統計情報
        st.markdown("### 📊 統計情報")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_holiday_hours = sum(sum(
                time_data['holiday_hours'] for time_data in data.values()
            ) for data in holiday_data.values())
            st.metric("総休日時間", f"{total_holiday_hours:.1f}時間")
        
        with col2:
            total_weekday_hours = sum(sum(
                time_data['weekday_hours'] for time_data in data.values()
            ) for data in holiday_data.values())
            st.metric("総平日時間", f"{total_weekday_hours:.1f}時間")
        
        with col3:
            total_hours = total_holiday_hours + total_weekday_hours
            st.metric("総時間", f"{total_hours:.1f}時間")
        
        with col4:
            holiday_ratio = (total_holiday_hours / total_hours * 100) if total_hours > 0 else 0
            st.metric("休日比率", f"{holiday_ratio:.1f}%")

def format_hours(hours):
    """時間を表示用の形式に変換する"""
    if hours == 0:
        return ""
    
    h = int(hours)
    m = int((hours - h) * 60)
    return f"{h}:{m:02d}"

def hours_to_decimal(hours):
    """時間を小数形式に変換する（1:30 → 1.5）"""
    if hours == 0:
        return 0
    
    h = int(hours)
    m = int((hours - h) * 60)
    return h + m / 60

def read_overtime_sheet(workbook):
    """残業代シートからメンバー名と単価を読み込む"""
    if "残業代" not in workbook.sheetnames:
        return {}
    
    worksheet = workbook["残業代"]
    member_data = {}
    
    # C30から空白セルが来るまで読み込み
    row = 30
    while True:
        cell_c = f"C{row}"
        member_name = worksheet[cell_c].value
        
        if member_name is None or str(member_name).strip() == "":
            break
        
        # D〜G列の単価を取得
        cell_d = f"D{row}"
        cell_e = f"E{row}"
        cell_f = f"F{row}"
        cell_g = f"G{row}"
        
        rate_d = worksheet[cell_d].value or 0
        rate_e = worksheet[cell_e].value or 0
        rate_f = worksheet[cell_f].value or 0
        rate_g = worksheet[cell_g].value or 0
        
        member_data[str(member_name).strip()] = {
            'D': float(rate_d) if rate_d else 0,
            'E': float(rate_e) if rate_e else 0,
            'F': float(rate_f) if rate_f else 0,
            'G': float(rate_g) if rate_g else 0
        }
        
        row += 1
    
    return member_data

def match_member_name(full_name, sheet_names):
    """フルネームとシート名を照合する"""
    for sheet_name in sheet_names:
        if sheet_name in full_name or full_name in sheet_name:
            return sheet_name
    return None

def calculate_overtime_pay(holiday_data, overtime_rates):
    """残業代を計算する"""
    pay_data = {}
    
    for member, data in holiday_data.items():
        # メンバー名とシート名の照合
        matched_sheet = None
        for full_name, rates in overtime_rates.items():
            if match_member_name(full_name, [member]):
                matched_sheet = member
                member_rates = rates
                break
        
        if not matched_sheet:
            continue
        
        member_pay = {}
        
        # 各時間帯の残業代を計算
        time_slots = [
            '休日時間帯の応動（09:00-18:00）',
            '平日・休日時間外の応動（18:00-22:00）',
            '平日・休日深夜の応動（22:00-05:00）',
            '平日・休日時間外の応動（05:00-09:00）'
        ]
        
        for time_slot in time_slots:
            if time_slot in data:
                time_data = data[time_slot]
                holiday_hours = hours_to_decimal(time_data['holiday_hours'])
                weekday_hours = hours_to_decimal(time_data['weekday_hours'])
                
                # 単価の組み合わせで計算
                if time_slot == '休日時間帯の応動（09:00-18:00）':
                    # 休日時間帯の応動（09:00-18:00）休日*F列
                    holiday_pay = holiday_hours * member_rates['F']
                    weekday_pay = 0  # この時間帯は平日なし
                elif time_slot == '平日・休日時間外の応動（18:00-22:00）':
                    # 平日・休日時間外の応動（18:00-22:00）休日*F列
                    # 平日・休日時間外の応動（18:00-22:00）平日*D列
                    holiday_pay = holiday_hours * member_rates['F']
                    weekday_pay = weekday_hours * member_rates['D']
                elif time_slot == '平日・休日深夜の応動（22:00-05:00）':
                    # 平日・休日深夜の応動（22:00-05:00）休日*G列
                    # 平日・休日深夜の応動（22:00-05:00）平日*E列
                    holiday_pay = holiday_hours * member_rates['G']
                    weekday_pay = weekday_hours * member_rates['E']
                elif time_slot == '平日・休日時間外の応動（05:00-09:00）':
                    # 平日・休日時間外の応動（05:00-09:00）休日*G列
                    # 平日・休日時間外の応動（05:00-09:00）平日*E列
                    holiday_pay = holiday_hours * member_rates['G']
                    weekday_pay = weekday_hours * member_rates['E']
                else:
                    holiday_pay = 0
                    weekday_pay = 0
                
                member_pay[time_slot] = {
                    'holiday_pay': holiday_pay,
                    'weekday_pay': weekday_pay,
                    'total_pay': holiday_pay + weekday_pay
                }
        
        pay_data[member] = member_pay
    
    return pay_data

def display_overtime_pay_results(pay_data, holiday_data):
    """残業代計算結果を表示する"""
    st.markdown("## 💰 残業代計算結果")
    
    # データフレームを作成（指定された形式）
    df_data = []
    for member, data in pay_data.items():
        row = {'メンバー': member}
        
        # 稼働時間と請求額を計算
        time_slots = [
            '休日時間帯の応動（09:00-18:00）',
            '平日・休日時間外の応動（18:00-22:00）',
            '平日・休日深夜の応動（22:00-05:00）',
            '平日・休日時間外の応動（05:00-09:00）'
        ]
        
        total_work_hours = 0
        total_pay = 0
        
        for time_slot in time_slots:
            if time_slot in data and time_slot in holiday_data[member]:
                time_data = data[time_slot]
                holiday_time_data = holiday_data[member][time_slot]
                
                # 稼働時間（休日+平日の合計時間）
                work_hours = hours_to_decimal(holiday_time_data['holiday_hours']) + hours_to_decimal(holiday_time_data['weekday_hours'])
                total_work_hours += work_hours
                
                # 請求額（休日+平日の合計金額）
                pay_amount = time_data['holiday_pay'] + time_data['weekday_pay']
                total_pay += pay_amount
                
                # 稼働時間列（0.0の場合は空白）
                if work_hours > 0:
                    row[f'稼働：{time_slot}'] = f"{work_hours:.1f}"
                else:
                    row[f'稼働：{time_slot}'] = ""
                
                # 請求額列（¥0の場合は空白）
                if pay_amount > 0:
                    row[f'請求：{time_slot}'] = f"¥{pay_amount:,.0f}"
                else:
                    row[f'請求：{time_slot}'] = ""
            else:
                row[f'稼働：{time_slot}'] = ""
                row[f'請求：{time_slot}'] = ""
        
        # 総稼働時間と総請求額（0の場合は空白）
        if total_work_hours > 0:
            row['稼働時間'] = f"{total_work_hours:.1f}"
        else:
            row['稼働時間'] = ""
        
        if total_pay > 0:
            row['請求額'] = f"¥{total_pay:,.0f}"
        else:
            row['請求額'] = ""
        
        df_data.append(row)
    
    if df_data:
        df = pd.DataFrame(df_data)
        
        # 列の順序を指定（稼働4つ左側、請求4つ右側）
        columns_order = [
            'メンバー',
            '稼働：休日時間帯の応動（09:00-18:00）',
            '稼働：平日・休日時間外の応動（18:00-22:00）',
            '稼働：平日・休日深夜の応動（22:00-05:00）',
            '稼働：平日・休日時間外の応動（05:00-09:00）',
            '請求：休日時間帯の応動（09:00-18:00）',
            '請求：平日・休日時間外の応動（18:00-22:00）',
            '請求：平日・休日深夜の応動（22:00-05:00）',
            '請求：平日・休日時間外の応動（05:00-09:00）',
            '稼働時間',
            '請求額'
        ]
        
        # 列の順序を適用
        df = df[columns_order]
        
        # 表示
        st.dataframe(df, use_container_width=True)
        
        # ダウンロードボタン
        csv = df.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📥 CSVファイルとしてダウンロード",
            data=csv,
            file_name=f"残業代計算_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
        
        # 統計情報
        st.markdown("### 📊 統計情報")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            total_pay = sum(sum(
                time_data['total_pay'] for time_data in data.values()
            ) for data in pay_data.values())
            st.metric("総請求額", f"¥{total_pay:,.0f}")
        
        with col2:
            total_hours = sum(sum(
                hours_to_decimal(holiday_time_data['holiday_hours']) + hours_to_decimal(holiday_time_data['weekday_hours'])
                for holiday_time_data in member_data.values()
            ) for member_data in holiday_data.values())
            st.metric("総稼働時間", f"{total_hours:.1f}")
        
        with col3:
            avg_pay = total_pay / len(pay_data) if pay_data else 0
            st.metric("平均請求額", f"¥{avg_pay:,.0f}")

def display_results(overtime_data):
    """結果を表示する"""
    st.markdown("## 📈 残業時間集計結果")
    
    # データフレームを作成
    df_data = []
    for member, data in overtime_data.items():
        row = {'メンバー': member}
        for time_slot, time_data in data.items():
            if isinstance(time_data, dict):
                row[time_slot] = time_data['display']
            else:
                row[time_slot] = time_data
        df_data.append(row)
    
    if df_data:
        df = pd.DataFrame(df_data)
        
        # 列の順序を指定
        columns_order = ['メンバー'] + list(df.columns[1:])
        df = df[columns_order]
        
        # 表示
        st.dataframe(df, use_container_width=True)
        
        # ダウンロードボタン
        csv = df.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📥 CSVファイルとしてダウンロード",
            data=csv,
            file_name=f"残業時間集計_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
        
        # 統計情報
        st.markdown("### 📊 統計情報")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("対象メンバー数", len(overtime_data))
        
        with col2:
            total_hours = sum(sum(
                time_data['hours'] if isinstance(time_data, dict) else time_data 
                for time_data in data.values()
            ) for data in overtime_data.values())
            st.metric("総残業時間", f"{total_hours:.1f}時間")
        
        with col3:
            # データがあるメンバーのみで平均を計算
            members_with_data = [data for data in overtime_data.values() if any(
                (time_data['hours'] if isinstance(time_data, dict) else time_data) > 0 
                for time_data in data.values()
            )]
            avg_hours = total_hours / len(members_with_data) if members_with_data else 0
            st.metric("平均残業時間", f"{avg_hours:.1f}時間")
        
        with col4:
            max_hours = max(sum(
                time_data['hours'] if isinstance(time_data, dict) else time_data 
                for time_data in data.values()
            ) for data in overtime_data.values()) if overtime_data else 0
            st.metric("最大残業時間", f"{max_hours:.1f}時間")

if __name__ == "__main__":
    main()
